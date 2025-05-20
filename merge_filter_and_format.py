import os
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Configurable Filters
FILTER_SEVERITY = ["Informational", "Low"]
FILTER_POLICY_IDS = ["XYZ-00123", "ABC-99999"]
OUTPUT_FILE = "merged_filtered_output.xlsx"

def list_excel_files():
    return [f for f in os.listdir() if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]

def check_column_consistency(files):
    ref_cols, _ = pd.read_excel(files[0], nrows=1), None
    mismatches = []
    for file in files[1:]:
        cols = pd.read_excel(file, nrows=1).columns
        if list(cols) != list(ref_cols.columns):
            mismatches.append(file)
    return len(mismatches) == 0, ref_cols.columns.tolist(), mismatches

def merge_files(files, columns):
    df_all = pd.DataFrame(columns=columns)
    for f in files:
        df = pd.read_excel(f)
        df_all = pd.concat([df_all, df], ignore_index=True)
        print(f"✅ Merged: {f} ({len(df)} rows)")
    return df_all

def filter_rows(df):
    original = len(df)

    if "Severity" in df.columns:
        df["Severity"] = df["Severity"].astype(str).str.strip()
        df = df[~df["Severity"].isin(FILTER_SEVERITY)]

    if "Policy ID" in df.columns:
        df["Policy ID"] = df["Policy ID"].astype(str).str.strip()
        df = df[~df["Policy ID"].isin([str(x).strip() for x in FILTER_POLICY_IDS])]

    print(f"🧹 Filtered out {original - len(df)} row(s)")
    return df

def format_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    ws.freeze_panes = "A2"
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_width = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        for cell in col:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.column_dimensions[col_letter].width = max_width + 2

    wb.save(filename)

def add_summary(df, path, prefix="Filtered_"):
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        if "Severity" in df.columns:
            summary = df["Severity"].value_counts().reset_index()
            summary.columns = ["Severity", "Count"]
            summary.loc[len(summary)] = ["Total Findings", len(df)]
            summary.to_excel(writer, sheet_name=f"{prefix}Summary_Overall", index=False)

        if "BU" in df.columns and "Severity" in df.columns:
            bu_summary = df.groupby(["BU", "Severity"]).size().unstack(fill_value=0).reset_index()
            bu_summary["Total"] = bu_summary.iloc[:, 1:].sum(axis=1)
            bu_summary.to_excel(writer, sheet_name=f"{prefix}Summary_By_BU", index=False)

def main():
    start = time.time()
    files = list_excel_files()
    if not files:
        print("⚠️ No Excel files found.")
        return

    valid, columns, mismatches = check_column_consistency(files)
    if not valid:
        print("❌ Column mismatch in files:", mismatches)
        return

    df = merge_files(files, columns)
    df = filter_rows(df)
    df.to_excel(OUTPUT_FILE, index=False)
    format_excel(OUTPUT_FILE)
    add_summary(df, OUTPUT_FILE)
    print(f"\n✅ Output written to: {OUTPUT_FILE}")
    print(f"⏱️ Done in {time.time() - start:.2f} seconds.")

if __name__ == "__main__":
    main()
