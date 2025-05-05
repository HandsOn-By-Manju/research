import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time

# Set input and output file names
input_csv = "input_file.csv"
output_excel = "output_file.xlsx"
anex_file = "Anex.xlsx"

# Columns to remove from the main file (if they exist)
columns_to_remove = [
    "DummyColumn1", "DummyColumn2", "DummyColumn3",
    "DummyColumn4", "DummyColumn5", "DummyColumn6",
    "DummyColumn7", "DummyColumn8", "DummyColumn9"
]

# New columns to add to the final output
columns_to_add = ["Col1", "Col2", "Col3", "Col4", "Col5", "Col6", "Col7", "Col8"]

start_time = time.time()
print("\n🚀 Starting the Excel processing script...")

# Step 1: Read the input CSV file
print(f"📄 Reading input CSV file: {input_csv}")
df = pd.read_csv(input_csv)
print("✅ Loaded input data.")

# Step 2: Extract 'ID' and 'Name' from 'Details' column if present
if 'Details' in df.columns:
    print("🔧 Extracting 'ID' and 'Name' from 'Details'")
    df['ID'] = df['Details'].str.extract(r'^\s*(\S+)\s*\(')[0].str.replace(r'\s+', '', regex=True)
    df['Name'] = df['Details'].str.extract(r'\((.*?)\)')[0].str.replace(r'\s+', '', regex=True)

# Step 3: Extract only filename from 'Res_ID' path
if 'Res_ID' in df.columns:
    print("🔧 Formatting 'Res_ID' to get only filenames")
    df['Res_ID'] = df['Res_ID'].apply(lambda x: str(x).split('/')[-1])

# Step 4: Remove any unwanted dummy columns
print("🧹 Dropping unwanted dummy columns (if present)")
df.drop(columns=[col for col in columns_to_remove if col in df.columns], inplace=True)

# Step 5: Add 8 empty columns to the DataFrame
print("➕ Adding empty columns: Col1 to Col8")
for col in columns_to_add:
    df[col] = ''

# Step 6: Merge policy data from Anex1 using 'Policy ID'
print("🔗 Mapping Policy Info from Anex1")
anex1 = pd.read_excel(anex_file, sheet_name="Anex1")[['Policy ID', 'Policy Statement', 'Policy Remediation']]
df = df.merge(anex1, on="Policy ID", how="left")
df['Col1'] = df['Policy Statement'].fillna("Policy details doesn't exist")
df['Col2'] = df['Policy Remediation'].fillna("Policy details doesn't exist")

# Log unmatched Policy IDs to file
unmatched_policies = df[df['Policy Statement'].isna() & df['Policy Remediation'].isna()]['Policy ID'].dropna().unique()
if len(unmatched_policies) > 0:
    print("⚠️ Unmatched Policy IDs:", unmatched_policies.tolist())
    with open("unmatched_policy_ids.txt", "w") as f:
        for pid in unmatched_policies:
            f.write(f"{pid}\n")
# Drop extra columns after mapping
df.drop(columns=['Policy Statement', 'Policy Remediation'], inplace=True, errors='ignore')

# Step 7: Merge subscription data from Anex2 using 'Subscription ID'
print("🔗 Mapping Subscription Info from Anex2")
anex2 = pd.read_excel(anex_file, sheet_name="Anex2")[['Subscription ID', 'Environment', 'Primary Contact']]
df = df.merge(anex2, on="Subscription ID", how="left")
df['Col3'] = df['Environment'].fillna("Environment/contact info not found")
df['Col4'] = df['Primary Contact'].fillna("Environment/contact info not found")

# Log unmatched Subscription IDs
unmatched_subs = df[df['Environment'].isna() & df['Primary Contact'].isna()]['Subscription ID'].dropna().unique()
if len(unmatched_subs) > 0:
    print("⚠️ Unmatched Subscription IDs:", unmatched_subs.tolist())
    with open("unmatched_subscription_ids.txt", "w") as f:
        for sid in unmatched_subs:
            f.write(f"{sid}\n")
df.drop(columns=['Environment', 'Primary Contact'], inplace=True, errors='ignore')

# Step 8: Merge M1-M4 data from Anex3 using 'Contact'
print("🔗 Mapping M1–M4 from Anex3 based on 'Contact'")
anex3 = pd.read_excel(anex_file, sheet_name="Anex3")[['Contact', 'M1', 'M2', 'M3', 'M4']]
df = df.merge(anex3, on="Contact", how="left")
df['Col5'] = df['M1'].fillna("Data not available")
df['Col6'] = df['M2'].fillna("Data not available")
df['Col7'] = df['M3'].fillna("Data not available")
df['Col8'] = df['M4'].fillna("Data not available")

# Log unmatched Contact values
unmatched_contacts = df[df[['M1', 'M2', 'M3', 'M4']].isna().all(axis=1)]['Contact'].dropna().unique()
if len(unmatched_contacts) > 0:
    print("⚠️ Unmatched Contact values:", unmatched_contacts.tolist())
    with open("unmatched_anex3_contacts.txt", "w") as f:
        for c in unmatched_contacts:
            f.write(f"{c}\n")
df.drop(columns=['M1', 'M2', 'M3', 'M4'], inplace=True, errors='ignore')

# Step 9: Save the updated DataFrame to Excel
print(f"💾 Writing final Excel to {output_excel}")
df.to_excel(output_excel, index=False)

# Step 10: Format Excel output (top-left align and rename sheet)
print("🎨 Formatting Excel: align top-left and rename sheet to 'Issues'")
wb = load_workbook(output_excel)
ws = wb.active
ws.title = "Issues"
alignment = Alignment(horizontal='left', vertical='top')
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = alignment
wb.save(output_excel)

# Final summary
print(f"✅ Script completed in {time.time() - start_time:.2f} seconds.")
