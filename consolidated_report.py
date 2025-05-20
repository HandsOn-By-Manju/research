import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# === START TIMER ===
start_time = time.time()

# === CONFIGURATION ===
folder_path = os.getcwd()
output_file = "merged_output.xlsx"
log_file = os.path.join(folder_path, "merge_log.txt")

# Configurable filters
remove_severity = ["Informational", "Low"]  # Severities to remove
remove_policy_ids = ["POL-001", "POL-999"]  # Policy IDs to remove

# === SAFETY CHECK ===
if not os.path.isdir(folder_path):
    print(f"❌ The folder does not exist: {folder_path}")
    exit()

# === INITIATE LOG ===
log_lines = []

# === STEP 1: Get All Excel Files ===
print(f"🔍 Current folder being scanned: {folder_path}")
print("📥 Searching for Excel files...")
excel_files = [f for f in os.listdir(folder_path) if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
log_lines.append(f"Number of Excel files found: {len(excel_files)}")
print(f"✅ Found {len(excel_files)} Excel files.")

if len(excel_files) == 0:
    log_lines.append("No Excel files found.")
    with open(log_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(log_lines))
    print("❌ No Excel files found in current folder.")
    exit()

# === STEP 2: Validate Column Names ===
print("🔎 Checking column structure of files...")
first_file = os.path.join(folder_path, excel_files[0])
first_df = pd.read_excel(first_file)
expected_columns = list(first_df.columns)
log_lines.append(f"Expected columns: {expected_columns}")
print(f"📌 Expected columns: {expected_columns}")

all_dataframes = []
row_counts = {}
severity_summary = {}
bu_severity_summary = {}

for file in excel_files:
    print(f"\n🔍 Reading file: {file}")
    path = os.path.join(folder_path, file)
    df = pd.read_excel(path)

    if list(df.columns) != expected_columns:
        log_lines.append(f"❌ Column mismatch in file: {file}")
        log_lines.append(f"Found columns: {list(df.columns)}")
        with open(log_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(log_lines))
        print(f"❌ Column mismatch found in '{file}'. Check merge_log.txt for details.")
        exit()

    original_rows = len(df)

    # === Apply filters ===
    if 'Severity' in df.columns:
        df = df[~df['Severity'].isin(remove_severity)]
    if 'Policy ID' in df.columns:
        df = df[~df['Policy ID'].isin(remove_policy_ids)]

    row_counts[file] = len(df)

    # === Per-file severity count ===
    if 'Severity' in df.columns:
        severity_count = df['Severity'].value_counts().to_dict()
        severity_summary[file] = severity_count
        log_lines.append(f"\nSeverity breakdown for {file}:")
        for sev, count in severity_count.items():
            log_lines.append(f"  {sev}: {count}")

    # === Business Unit-wise severity count ===
    if 'Business Unit' in df.columns and 'Severity' in df.columns:
        bu_group = df.groupby(['Business Unit', 'Severity']).size().reset_index(name='Count')
        for _, row in bu_group.iterrows():
            bu = row['Business Unit']
            sev = row['Severity']
            cnt = row['Count']
            bu_severity_summary.setdefault(bu, {}).setdefault(sev, 0)
            bu_severity_summary[bu][sev] += cnt

    all_dataframes.append(df)
    print(f"✅ File '{file}': Loaded {original_rows} → Retained {len(df)} rows after filtering.")

# === STEP 3: Merge Files ===
print("\n🔧 Merging all data...")
merged_df = pd.concat(all_dataframes, ignore_index=True)
output_path = os.path.join(folder_path, output_file)
merged_df.to_excel(output_path, index=False)
print(f"✅ Merged data written to '{output_file}'")

# === STEP 4: Excel Formatting ===
print("🎨 Applying Excel formatting...")
wb = load_workbook(output_path)
ws = wb.active

header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill

for col in ws.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_length + 2

ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = "A2"
wb.save(output_path)
print("✅ Excel formatting complete.")

# === STEP 5: Write Log File ===
log_lines.append("\nRow counts per file (after filtering):")
for file, count in row_counts.items():
    log_lines.append(f"{file}: {count} rows")
log_lines.append(f"\nTotal rows in merged file: {len(merged_df)}")

# === Business Unit Summary ===
if bu_severity_summary:
    log_lines.append("\nBusiness Unit-wise Severity Summary:")
    for bu, sev_dict in bu_severity_summary.items():
        log_lines.append(f"  {bu}:")
        for sev, cnt in sev_dict.items():
            log_lines.append(f"    {sev}: {cnt}")

# === STEP 6: Show Time Taken in Smart Format ===
end_time = time.time()
elapsed = int(end_time - start_time)

def format_time(seconds):
    if seconds < 60:
        return f"{seconds} seconds"
    elif seconds < 3600:
        minutes = seconds // 60
        sec = seconds % 60
        return f"{minutes} minutes {sec} seconds"
    else:
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        sec = seconds % 60
        return f"{hours} hours {minutes} minutes {sec} seconds"

formatted_time = format_time(elapsed)
log_lines.append(f"\n⏱️ Time taken: {formatted_time}")

with open(log_file, "w", encoding="utf-8") as f:
    f.write("\n".join(log_lines))

# === FINAL OUTPUT ===
print("\n📄 Log written to merge_log.txt")
print("\n📊 Summary:")
for file, count in row_counts.items():
    print(f"   - {file}: {count} rows")
print(f"\n📈 Total rows merged: {len(merged_df)}")
print(f"⏱️ Time taken to complete: {formatted_time}")
print(f"✅ Merged Excel file: {output_path}")
print(f"📝 Log file created at: {log_file}")
