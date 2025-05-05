import pandas as pd
import time
import difflib
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# === Configurable Inputs ===
input_csv = "input_file.csv"
output_excel = "output_step5.xlsx"
anex_file = "Report_Anex.xlsx"

account_column_name = "Account"
resource_column_name = "Resource ID"
parse_account_column = True

anex1_sheet = "Anex1_Remediation_Sheet"
anex2_sheet = "Anex2_Sub_Sheet"
anex3_sheet = "Anex3_Contact_Sheet"

primary_contact_column = "Primary Contact"
manager_columns = [
    "Manager / Sr Manager / Director / Sr Director",
    "Sr Director / VP",
    "VP / SVP / CVP",
    "BU"
]

columns_to_remove = [
    "DummyColumn1", "DummyColumn2", "DummyColumn3",
    "DummyColumn4", "DummyColumn5", "DummyColumn6",
    "DummyColumn7", "DummyColumn8", "DummyColumn9"
]

columns_to_add = [
    "Description",
    "Remediation Steps",
    "Environment",
    primary_contact_column,
] + manager_columns

validation_checks = [
    {
        "name": "Subscription ID",
        "sheet": anex2_sheet,
        "join_column": "Subscription ID"
    },
    {
        "name": "Policy ID",
        "sheet": anex1_sheet,
        "join_column": "Policy ID"
    }
]

# === Start Execution ===
start_time = time.time()
print("\n🚀 Starting preprocessing and validation...")

# Step 1: Load CSV
df = pd.read_csv(input_csv)
print(f"✅ Loaded input file: {input_csv}")

# Step 2: Parse Subscription ID and Name
if parse_account_column and account_column_name in df.columns:
    print(f"🔧 Extracting Subscription ID and Name from '{account_column_name}'")
    df["Subscription ID"] = df[account_column_name].str.extract(r"^(\S+)\s*\(")[0].str.replace(r"\s+", "", regex=True)
    df["Subscription Name"] = df[account_column_name].str.extract(r"\((.*?)\)")[0].str.replace(r"\s+", "", regex=True)

# Step 3: Clean Resource ID
if resource_column_name in df.columns:
    print(f"🔧 Cleaning '{resource_column_name}' to extract filename")
    df[resource_column_name] = df[resource_column_name].apply(lambda x: str(x).split("/")[-1])

# Step 4: Drop unwanted columns
existing_to_drop = [col for col in columns_to_remove if col in df.columns]
df.drop(columns=existing_to_drop, inplace=True)
print(f"🧹 Dropped columns: {existing_to_drop if existing_to_drop else 'None'}")

# Step 5: Add empty columns
print(f"➕ Adding columns: {columns_to_add}")
for col in columns_to_add:
    df[col] = ""

# Step 6: Validate Subscription ID and Policy ID
for check in validation_checks:
    name = check["name"]
    sheet = check["sheet"]
    join_column = check["join_column"]

    print(f"\n🔍 Validating {name} using sheet '{sheet}'")
    try:
        df[join_column] = df[join_column].astype(str).str.strip()
        df_anex = pd.read_excel(anex_file, sheet_name=sheet)
        df_anex[join_column] = df_anex[join_column].astype(str).str.strip()

        input_ids = set(df[join_column].dropna())
        anex_ids = set(df_anex[join_column].dropna())

        unmatched_ids = sorted(input_ids - anex_ids)
        matched_ids = sorted(input_ids & anex_ids)

        if unmatched_ids:
            with open(f"unmatched_{join_column.replace(' ', '_').lower()}.txt", "w") as f:
                for val in unmatched_ids:
                    f.write(f"{val}\n")
            print(f"❌ {len(unmatched_ids)} unmatched {name}(s) saved to file.")
        else:
            print(f"✅ All {name}s matched!")

        print(f"📊 {name} Summary:")
        print(f"   Total     : {len(input_ids)}")
        print(f"   Matched   : {len(matched_ids)}")
        print(f"   Unmatched : {len(unmatched_ids)}")
        print(f"   Match %   : {round((len(matched_ids)/len(input_ids))*100, 2)}%")

    except Exception as e:
        print(f"❌ Error validating {name}: {e}")

# Step 7: Map Description and Remediation Steps
print("\n🧩 Mapping 'Description' and 'Remediation Steps'")
try:
    df_remed = pd.read_excel(anex_file, sheet_name=anex1_sheet)
    df_remed["Policy ID"] = df_remed["Policy ID"].astype(str).str.strip()

    df = df.merge(df_remed[["Policy ID", "Policy Statement", "Policy Remediation"]],
                  on="Policy ID", how="left")

    df["Description"] = df["Policy Statement"].fillna("Policy details not available")
    df["Remediation Steps"] = df["Policy Remediation"].fillna("Remediation steps not available")

    df.drop(columns=["Policy Statement", "Policy Remediation"], inplace=True)
    print("✅ Remediation fields mapped successfully.")
except Exception as e:
    print(f"❌ Error mapping remediation data: {e}")

# Step 8: Map Environment and Primary Contact
print("\n📌 Mapping 'Environment' and 'Primary Contact'")
try:
    df_env = pd.read_excel(anex_file, sheet_name=anex2_sheet)
    df_env["Subscription ID"] = df_env["Subscription ID"].astype(str).str.strip()

    df.drop(columns=["Environment", primary_contact_column], inplace=True, errors="ignore")
    df = df.merge(df_env[["Subscription ID", "Environment", primary_contact_column]],
                  on="Subscription ID", how="left")

    df["Environment"] = df["Environment"].fillna("Environment not available")
    df[primary_contact_column] = df[primary_contact_column].fillna("Primary contact not available")

    print("✅ Environment and contact data filled.")
except Exception as e:
    print(f"❌ Error mapping environment/contact: {e}")

# Step 9: Validate and Map Manager Hierarchy by Primary Contact
print("\n📌 Validating and mapping Manager/VP/BU from Primary Contact")
try:
    df_contact = pd.read_excel(anex_file, sheet_name=anex3_sheet)
    df_contact.columns = df_contact.columns.str.strip()
    df.columns = df.columns.str.strip()

    actual_columns = df_contact.columns.tolist()
    missing_columns = [col for col in manager_columns if col not in actual_columns]

    if missing_columns:
        print(f"❌ Missing columns in '{anex3_sheet}': {missing_columns}")
        for col in missing_columns:
            suggestions = difflib.get_close_matches(col, actual_columns, n=1, cutoff=0.6)
            if suggestions:
                print(f"   💡 Did you mean: {suggestions[0]} for '{col}'?")
        raise Exception("Required columns not found in contact sheet")

    df_contact[primary_contact_column] = df_contact[primary_contact_column].astype(str).str.strip()
    df[primary_contact_column] = df[primary_contact_column].astype(str).str.strip()

    input_contacts = set(df[primary_contact_column].dropna())
    anex_contacts = set(df_contact[primary_contact_column].dropna())

    unmatched = sorted(input_contacts - anex_contacts)
    matched = sorted(input_contacts & anex_contacts)

    if unmatched:
        with open("unmatched_primary_contact.txt", "w") as f:
            for val in unmatched:
                f.write(f"{val}\n")
        print(f"❌ {len(unmatched)} unmatched Primary Contact(s) saved to file.")
    else:
        print("✅ All Primary Contacts matched!")

    print(f"📊 Primary Contact Summary:")
    print(f"   Total     : {len(input_contacts)}")
    print(f"   Matched   : {len(matched)}")
    print(f"   Unmatched : {len(unmatched)}")
    print(f"   Match %   : {round((len(matched)/len(input_contacts))*100, 2)}%")

    df.drop(columns=manager_columns, inplace=True, errors="ignore")
    df = df.merge(df_contact[[primary_contact_column] + manager_columns], on=primary_contact_column, how="left")

    print("✅ Manager hierarchy and BU data filled.")

except Exception as e:
    print(f"❌ Error during contact mapping: {e}")

# Step 10: Save Output to Excel with Formatting
print("\n💾 Saving Excel file with formatting...")
try:
    df.to_excel(output_excel, index=False)

    wb = load_workbook(output_excel)
    ws = wb.active

    align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align

    ws.freeze_panes = 'A2'

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    wb.save(output_excel)
    print(f"✅ Final file saved with alignment and formatting: {output_excel}")
except Exception as e:
    print(f"❌ Error applying formatting: {e}")

print(f"⏱️ Total time: {time.time() - start_time:.2f} seconds")
