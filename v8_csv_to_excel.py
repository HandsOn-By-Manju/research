import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Step 1: Read main CSV file
csv_file_path = 'input_file.csv'  # Replace with your actual CSV file path
df = pd.read_csv(csv_file_path)

# Step 2: Show original columns
print("Original Columns:")
print(df.columns.tolist())

# Step 3: Extract 'ID' and 'Name' from 'Details'
if 'Details' in df.columns:
    df['ID'] = df['Details'].str.extract(r'^\s*(\S+)\s*\(')[0].str.replace(r'\s+', '', regex=True)
    df['Name'] = df['Details'].str.extract(r'\((.*?)\)')[0].str.replace(r'\s+', '', regex=True)

# Step 4: Format 'Res_ID' to keep only filename from forward-slash path
if 'Res_ID' in df.columns:
    df['Res_ID'] = df['Res_ID'].apply(lambda x: str(x).split('/')[-1])

# Step 5: Define columns to remove
columns_to_remove = [
    'DummyColumn1',
    'DummyColumn2',
    'DummyColumn3',
    'DummyColumn4',
    'DummyColumn5',
    'DummyColumn6',
    'DummyColumn7',
    'DummyColumn8',
    'DummyColumn9'
]

# Drop columns only if they exist
existing_cols_to_remove = [col for col in columns_to_remove if col in df.columns]
df.drop(columns=existing_cols_to_remove, inplace=True)

# Step 6: Add 8 empty columns named Col1 to Col8
new_columns = ['Col1', 'Col2', 'Col3', 'Col4', 'Col5', 'Col6', 'Col7', 'Col8']
for col in new_columns:
    df[col] = ''

# Step 6.1: Fill Col1 and Col2 by matching Policy ID from Anex.xlsx (sheet Anex1)
anex_file = 'Anex.xlsx'
anex_sheet = 'Anex1'

# Read mapping sheet
anex_df = pd.read_excel(anex_file, sheet_name=anex_sheet)

# Keep relevant columns and drop duplicates
anex_df = anex_df[['Policy ID', 'Policy Statement', 'Policy Remediation']].drop_duplicates()

# Merge with main DataFrame on Policy ID
df = df.merge(anex_df, on='Policy ID', how='left')

# Fill Col1 and Col2 from mapping
df['Col1'] = df['Policy Statement']
df['Col2'] = df['Policy Remediation']

# Detect unmatched Policy IDs (where both Policy Statement and Remediation are missing)
unmatched_ids = df[df['Policy Statement'].isna() & df['Policy Remediation'].isna()]['Policy ID'].dropna().unique()

# Log to console
if len(unmatched_ids) > 0:
    print("\n⚠️ Unmatched Policy IDs:")
    for pid in unmatched_ids:
        print(f"- {pid}")
    # Save to file
    with open("unmatched_policy_ids.txt", "w") as f:
        for pid in unmatched_ids:
            f.write(f"{pid}\n")
    print("\n✅ Unmatched Policy IDs saved to 'unmatched_policy_ids.txt'")
else:
    print("\n✅ All Policy IDs matched successfully.")

# Replace NaNs with custom message for unmatched entries
message = "Policy details doesn't exist"
df['Col1'] = df['Col1'].fillna(message)
df['Col2'] = df['Col2'].fillna(message)

# Drop the now-unneeded merged columns
df.drop(columns=['Policy Statement', 'Policy Remediation'], inplace=True, errors='ignore')

# Step 7: Show updated columns
print("\nUpdated Columns:")
print(df.columns.tolist())

# Step 8: Save to Excel
excel_file_path = 'output_file.xlsx'
df.to_excel(excel_file_path, index=False)

# Step 9: Apply top-left alignment and rename sheet
wb = load_workbook(excel_file_path)
ws = wb.active
ws.title = "Issues"  # Rename the sheet to 'Issues'

alignment = Alignment(horizontal='left', vertical='top')
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = alignment

wb.save(excel_file_path)
print(f"\n✅ Excel file saved successfully with sheet name 'Issues': {excel_file_path}")
