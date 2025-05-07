import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Start timer
start_time = time.time()

# File paths
input_file_path = "input_file.csv"
remediation_file_path = "remediation_file.xlsx"
subscription_details_path = "subscription_details.xlsx"
ownership_file_path = "ownership_file.xlsx"
output_excel_path = "output_file.xlsx"

unmatched_policy_log_path = "unmatched_policy_ids.txt"
unmatched_subscription_log_path = "unmatched_subscription_ids.txt"
unmatched_primary_contact_log_path = "unmatched_primary_contacts.txt"

# Load input files
print("📥 Loading input files...")
df = pd.read_csv(input_file_path)
remediation_df = pd.read_excel(remediation_file_path)
subscription_df = pd.read_excel(subscription_details_path)
ownership_df = pd.read_excel(ownership_file_path)
print(f"🧾 Columns in input file: {list(df.columns)}")

# Rename 'Policy statement' to 'Description'
if 'Policy statement' in df.columns:
    df.rename(columns={'Policy statement': 'Description'}, inplace=True)
    print("✅ Renamed 'Policy statement' to 'Description'.")

# Extract Subscription ID and Name
if 'Account' in df.columns:
    print("🔍 Extracting 'Subscription ID' and 'Subscription Name' from 'Account'...")
    extracted = df['Account'].str.extract(r'([^()]+)\s*\(\s*([^()]+)\s*\)')
    extracted.columns = ['Subscription ID', 'Subscription Name']
    extracted['Subscription ID'] = extracted['Subscription ID'].str.replace(' ', '', regex=False)
    extracted['Subscription Name'] = extracted['Subscription Name'].str.replace(' ', '', regex=False)
    account_index = df.columns.get_loc('Account')
    for i, col in enumerate(extracted.columns):
        df.insert(account_index + 1 + i, col, extracted[col])
    print("✅ Added 'Subscription ID' and 'Subscription Name'.")

# Clean 'Resource ID'
if 'Resource ID' in df.columns:
    print("🧹 Cleaning 'Resource ID' column...")
    df['Resource ID'] = df['Resource ID'].astype(str).apply(
        lambda x: x.rstrip('/').rsplit('/', 1)[-1] if '/' in x.rstrip('/') else x
    )
    print("✅ Cleaned 'Resource ID'.")

# Validate and merge remediation data
print("🔍 Validating and merging 'Policy ID'...")
input_policy_ids = set(df['Policy ID'].dropna())
remediation_policy_ids = set(remediation_df['Policy ID'].dropna())
unmatched = input_policy_ids - remediation_policy_ids
if unmatched:
    with open(unmatched_policy_log_path, 'w') as f:
        for item in unmatched:
            f.write(f"{item}\n")
    print(f"⚠️ {len(unmatched)} unmatched Policy IDs written to {unmatched_policy_log_path}")
else:
    print("✅ All Policy IDs matched.")
df = df.merge(remediation_df[['Policy ID', 'Policy Statement', 'Policy Remediation']], on='Policy ID', how='left')
print("✅ Merged 'Policy Statement' and 'Policy Remediation'.")

# Validate and merge subscription details
print("🔍 Validating and merging 'Subscription ID'...")
input_sub_ids = set(df['Subscription ID'].dropna())
sub_ids = set(subscription_df['Subscription ID'].dropna())
unmatched = input_sub_ids - sub_ids
if unmatched:
    with open(unmatched_subscription_log_path, 'w') as f:
        for item in unmatched:
            f.write(f"{item}\n")
    print(f"⚠️ {len(unmatched)} unmatched Subscription IDs written to {unmatched_subscription_log_path}")
else:
    print("✅ All Subscription IDs matched.")
df = df.merge(subscription_df[['Subscription ID', 'Environment', 'BU', 'Primary Contact']], on='Subscription ID', how='left')
print("✅ Merged 'Environment', 'BU', and 'Primary Contact'.")

# Validate and merge ownership data
print("🔍 Validating and merging 'Primary Contact'...")
input_contacts = set(df['Primary Contact'].dropna())
ownership_contacts = set(ownership_df['Primary Contact'].dropna())
unmatched = input_contacts - ownership_contacts
if unmatched:
    with open(unmatched_primary_contact_log_path, 'w') as f:
        for item in unmatched:
            f.write(f"{item}\n")
    print(f"⚠️ {len(unmatched)} unmatched Primary Contacts written to {unmatched_primary_contact_log_path}")
else:
    print("✅ All Primary Contacts matched.")
df = df.merge(
    ownership_df[['Primary Contact', 'Manager / Sr Manager / Director / Sr Director', 'Sr Director / VP', 'VP / SVP / CVP']],
    on='Primary Contact', how='left'
)
print("✅ Merged ownership hierarchy columns.")

# Remove unwanted columns
columns_to_remove = ['Column1', 'Column2', 'Column3']
existing = [col for col in columns_to_remove if col in df.columns]
df.drop(columns=existing, inplace=True)
print(f"🧽 Dropped columns: {existing}")

# Rename standard columns
df.rename(columns={
    'Cloud provider': 'Cloud Provider',
    'Environment_y': 'Environment'
}, inplace=True)
print("🔤 Renamed columns: 'Cloud provider' → 'Cloud Provider', 'Environment_y' → 'Environment'")

# Reorder columns
desired_order = [
    'Cloud Provider', 'Subscription ID', 'Subscription Name', 'Region', 'Service', 'Resource ID',
    'Policy ID', 'Description', 'Severity', 'Policy Statement', 'Policy Remediation', 'Finding',
    'Environment', 'Primary Contact',
    'Manager / Sr Manager / Director / Sr Director', 'Sr Director / VP', 'VP / SVP / CVP', 'BU'
]
missing = [col for col in desired_order if col not in df.columns]
if missing:
    print(f"⚠️ Missing columns not included in output: {missing}")
df = df[[col for col in desired_order if col in df.columns]]
print("📐 Columns rearranged in specified order.")

# Save to Excel
df.to_excel(output_excel_path, index=False)
print(f"💾 Saved Excel before formatting: {output_excel_path}")

# Format using openpyxl
wb = load_workbook(output_excel_path)
ws = wb.active
ws.freeze_panes = "A2"

alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

# Apply header formatting
for cell in ws[1]:
    cell.fill = header_fill
    cell.alignment = alignment

# Apply alignment and column width
for col_idx, column_cells in enumerate(ws.columns, 1):
    max_len = 0
    for cell in column_cells:
        cell.alignment = alignment
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

wb.save(output_excel_path)
print("🎨 Applied formatting: word wrap, alignment, header color, column width, freeze header.")
print(f"✅ Final Excel file saved to: {output_excel_path}")

# Execution time
elapsed = time.time() - start_time
if elapsed < 60:
    print(f"\n⏱️ Execution Time: {elapsed:.2f} seconds")
elif elapsed < 3600:
    print(f"\n⏱️ Execution Time: {int(elapsed // 60)} minutes {elapsed % 60:.2f} seconds")
else:
    h, rem = divmod(elapsed, 3600)
    m, s = divmod(rem, 60)
    print(f"\n⏱️ Execution Time: {int(h)}h {int(m)}m {s:.2f}s")
