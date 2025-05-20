import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# === Configuration ===
excel_file_path = "your_excel_file.xlsx"               # <- Input Excel file
sheet_name = 0                                          # <- Sheet index or name
filter_column_name = "Severity"                        # <- Column to group by
target_column_name = "Subscription ID"                 # <- Column to extract

# Per-item prefix/suffix
prefix = "prefix_data_for_each_subscriptions"
suffix = "suffix_data_for_each_subscriptions"

# Base URL and final suffix
final_base_url = "https://example.com/trigger/"
final_suffix = "_FINAL_SUFFIX"

# Output file
output_excel_file = "Generated_URL_Report.xlsx"

# === Step 1: Load Excel ===
print("[INFO] Loading Excel file...")
df = pd.read_excel(excel_file_path, sheet_name=sheet_name)

# === Step 2: Validate Columns ===
missing = [col for col in [filter_column_name, target_column_name] if col not in df.columns]
if missing:
    print(f"[ERROR] Missing columns: {', '.join(missing)}")
    exit()

# === Step 3: Prepare Data ===
unique_filter_values = df[filter_column_name].dropna().astype(str).unique().tolist()
records = []

print(f"[INFO] Processing {len(unique_filter_values)} unique values from '{filter_column_name}'...\n")

for filter_val in unique_filter_values:
    filtered_df = df[df[filter_column_name].astype(str).str.strip().str.lower() ==
                     filter_val.strip().lower()]

    values = filtered_df[target_column_name].dropna().astype(str).unique().tolist()
    count = len(values)

    formatted_items = []
    for i, val in enumerate(values):
        item = f"{prefix}{val}"
        if i < len(values) - 1:
            item += suffix
        formatted_items.append(item)

    combined_string = "".join(formatted_items)
    full_url = f"{final_base_url}{combined_string}{final_suffix}"
    hyperlink_formula = f'=HYPERLINK("{full_url}", "Open URL")'

    records.append({
        filter_column_name: filter_val,
        "Subscription Count": count,
        "Open URL": hyperlink_formula,
        "Full URL (Debug)": full_url
    })

# === Step 4: Write to Excel with Formatting ===
print(f"[INFO] Writing results to '{output_excel_file}'...")

wb = Workbook()
ws = wb.active
ws.title = "URL Report"

headers = [filter_column_name, "Subscription Count", "Open URL", "Full URL (Debug)"]
ws.append(headers)

# Header formatting
header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data rows
for row in records:
    ws.append([
        row[filter_column_name],
        row["Subscription Count"],
        row["Open URL"],
        row["Full URL (Debug)"]
    ])

# Freeze top row
ws.freeze_panes = "A2"

# Auto column width + wrap text
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 5
    ws.column_dimensions[col_letter].width = adjusted_width

# Hide full URL debug column
ws.column_dimensions[get_column_letter(headers.index("Full URL (Debug)") + 1)].hidden = True

# Save file
wb.save(output_excel_file)
print("[SUCCESS] Excel file created with filters, formatting, and clickable links.\n")
