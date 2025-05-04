import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time

start_time = time.time()
print("🚀 Starting the Excel processing script...\n")

# Step 1: Read main CSV file
step_start = time.time()
csv_file_path = 'input_file.csv'
print(f"📄 Reading main CSV file: {csv_file_path}")
df = pd.read_csv(csv_file_path)
print(f"✅ Main CSV file loaded in {time.time() - step_start:.2f} seconds.")
print("🔍 Initial Columns:", df.columns.tolist())

# Step 2: Extract 'ID' and 'Name' from 'Details'
step_start = time.time()
print("\n🔧 Extracting ID and Name from 'Details' column (if exists)...")
if 'Details' in df.columns:
    df['ID'] = df['Details'].str.extract(r'^\s*(\S+)\s*\(')[0].str.replace(r'\s+', '', regex=True)
    df['Name'] = df['Details'].str.extract(r'\((.*?)\)')[0].str.replace(r'\s+', '', regex=True)
    print(f"✅ 'ID' and 'Name' columns extracted in {time.time() - step_start:.2f} seconds.")
else:
    print("⚠️ 'Details' column not found. Skipping ID and Name extraction.")

# Step 3: Format 'Res_ID'
step_start = time.time()
print("\n🔧 Formatting 'Res_ID' to keep only filename...")
if 'Res_ID' in df.columns:
    df['Res_ID'] = df['Res_ID'].apply(lambda x: str(x).split('/')[-1])
    print(f"✅ 'Res_ID' column formatted in {time.time() - step_start:.2f} seconds.")
else:
    print("⚠️ 'Res_ID' column not found. Skipping.")

# Step 4: Drop unwanted columns
step_start = time.time()
print("\n🧹 Dropping unused/dummy columns if present...")
columns_to_remove = [f'DummyColumn{i}' for i in range(1, 10)]
existing_cols_to_remove = [col for col in columns_to_remove if col in df.columns]
df.drop(columns=existing_cols_to_remove, inplace=True)
print(f"✅ Dropped columns: {existing_cols_to_remove if existing_cols_to_remove else 'None'} in {time.time() - step_start:.2f} seconds.")

# Step 5: Add 8 new columns Col1 to Col8
step_start = time.time()
print("\n➕ Adding empty columns: Col1 to Col8...")
new_columns = [f'Col{i}' for i in range(1, 9)]
for col in new_columns:
    df[col] = ''
print(f"✅ Empty columns added in {time.time() - step_start:.2f} seconds.")

# Step 6: Fill Col1 and Col2 using Anex1
step_start = time.time()
anex_file = 'Anex.xlsx'
print(f"\n📄 Reading Policy mapping sheet 'Anex1' from {anex_file}...")
anex_df1 = pd.read_excel(anex_file, sheet_name='Anex1')
anex_df1 = anex_df1[['Policy ID', 'Policy Statement', 'Policy Remediation']]

print("🔗 Merging on 'Policy ID'...")
df = df.merge(anex_df1, on='Policy ID', how='left')
df['Col1'] = df['Policy Statement']
df['Col2'] = df['Policy Remediation']

# Log unmatched Policy IDs
unmatched_ids = df[df['Policy Statement'].isna() & df['Policy Remediation'].isna()]['Policy ID'].dropna().unique()
if len(unmatched_ids) > 0:
    print("\n⚠️ Unmatched Policy IDs:")
    for pid in unmatched_ids:
        print(f"- {pid}")
    with open("unmatched_policy_ids.txt", "w") as f:
        for pid in unmatched_ids:
            f.write(f"{pid}\n")
    print("📁 Unmatched Policy IDs saved to 'unmatched_policy_ids.txt'")
else:
    print("✅ All Policy IDs matched successfully.")

# Fill unmatched
policy_msg = "Policy details doesn't exist"
df['Col1'] = df['Col1'].fillna(policy_msg)
df['Col2'] = df['Col2'].fillna(policy_msg)
df.drop(columns=['Policy Statement', 'Policy Remediation'], inplace=True, errors='ignore')
print(f"✅ Col1 & Col2 updated from 'Anex1' in {time.time() - step_start:.2f} seconds.")

# Step 7: Fill Col3 and Col4 using Anex2
step_start = time.time()
print(f"\n📄 Reading Subscription mapping sheet 'Anex2' from {anex_file}...")
anex_df2 = pd.read_excel(anex_file, sheet_name='Anex2')
anex_df2 = anex_df2[['Subscription ID', 'Environment', 'Primary Contact']]

print("🔗 Merging on 'Subscription ID'...")
df = df.merge(anex_df2, on='Subscription ID', how='left')
df['Col3'] = df['Environment']
df['Col4'] = df['Primary Contact']

# Log unmatched Subscription IDs
unmatched_subs = df[df['Environment'].isna() & df['Primary Contact'].isna()]['Subscription ID'].dropna().unique()
if len(unmatched_subs) > 0:
    print("\n⚠️ Unmatched Subscription IDs:")
    for sub in unmatched_subs:
        print(f"- {sub}")
    with open("unmatched_subscription_ids.txt", "w") as f:
        for sub in unmatched_subs:
            f.write(f"{sub}\n")
    print("📁 Unmatched Subscription IDs saved to 'unmatched_subscription_ids.txt'")
else:
    print("✅ All Subscription IDs matched successfully.")

# Fill unmatched
env_msg = "Environment/contact info not found"
df['Col3'] = df['Col3'].fillna(env_msg)
df['Col4'] = df['Col4'].fillna(env_msg)
df.drop(columns=['Environment', 'Primary Contact'], inplace=True, errors='ignore')
print(f"✅ Col3 & Col4 updated from 'Anex2' in {time.time() - step_start:.2f} seconds.")

# Step 8: Save to Excel
step_start = time.time()
excel_file_path = 'output_file.xlsx'
print(f"\n💾 Saving final Excel file: {excel_file_path}")
df.to_excel(excel_file_path, index=False)
print(f"✅ Excel file written in {time.time() - step_start:.2f} seconds.")

# Step 9: Apply formatting and rename sheet
step_start = time.time()
print("\n🎨 Applying top-left alignment and renaming sheet to 'Issues'...")
wb = load_workbook(excel_file_path)
ws = wb.active
ws.title = "Issues"
alignment = Alignment(horizontal='left', vertical='top')
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = alignment
wb.save(excel_file_path)
print(f"✅ Formatting applied and saved in {time.time() - step_start:.2f} seconds.")

# Done!
total_time = time.time() - start_time
print(f"\n✅ Script completed in {total_time:.2f} seconds. Output: {excel_file_path}")
