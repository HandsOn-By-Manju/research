import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# === Configurable Inputs ===
input_csv = "input_file.csv"
output_excel = "output_step5.xlsx"

remediation_file = "Remediation_Master_Sheet.xlsx"
subscription_file = "Sub_Data_file.xlsx"
ownership_file = "Ownership.xlsx"

account_column_name = "Account"
resource_column_name = "Affected Resource"
parse_account_column = True

primary_contact_column = "Primary Contact"
manager_columns = [
    "Manager / Sr Manager / Director / Sr Director",
    "Sr Director / VP",
    "VP / SVP / CVP",
    "BU"
]

columns_to_remove = [
    "DummyColumn1", "DummyColumn2", "DummyColumn3",
    "DummyColumn4", "DummyColumn5", "DummyColumn6",
    "DummyColumn7", "DummyColumn8", "DummyColumn9"
]

columns_to_add = [
    "Description", "Remediation Steps", "Environment", primary_contact_column
] + manager_columns

final_columns = [
    "Cloud Provider", "Subscription ID", "Subscription Name",
    "Policy ID", "Policy Statement", "Affected Resource",
    "Severity", "Description", "Remediation Steps",
    "Region", "Service", "Environment", "Primary Contact",
    "Manager / Sr Manager / Director / Sr Director", "Sr Director / VP",
    "VP / SVP / CVP", "BU", "Account", "Finding"
]

# === Utility Functions ===
def validate_required_columns(df, required_cols, source_name):
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        print(f"❌ Missing columns in {source_name}: {missing}")
        raise Exception(f"Missing columns in {source_name}")
    else:
        print(f"✅ All required columns present in {source_name}")

def format_duration(seconds):
    mins, secs = divmod(seconds, 60)
    hrs, mins = divmod(mins, 60)
    if hrs:
        return f"{int(hrs)}h {int(mins)}m {secs:.2f}s"
    elif mins:
        return f"{int(mins)}m {secs:.2f}s"
    else:
        return f"{secs:.2f}s"

# === Start Script ===
start_time = time.time()
print("\n🚀 Starting preprocessing and validation...")

# Step 1: Load CSV
df = pd.read_csv(input_csv)
print(f"✅ Loaded input file: {input_csv}")

# Step 2: Normalize known columns
df.columns = df.columns.str.strip()
df.rename(columns={
    "Cloud provider": "Cloud Provider",
    "Policy statement": "Policy Statement",
    "Resource ID": "Affected Resource"
}, inplace=True)

# Step 3: Parse Account for Subscription ID & Name
if parse_account_column and account_column_name in df.columns:
    print(f"🔧 Extracting Subscription ID and Name from '{account_column_name}'")
    df["Subscription ID"] = df[account_column_name].str.extract(r"^(\S+)\s*\(")[0].str.replace(r"\s+", "", regex=True)
    df["Subscription Name"] = df[account_column_name].str.extract(r"\((.*?)\)")[0].str.replace(r"\s+", "", regex=True)

# Step 4: Clean Affected Resource filename
if resource_column_name in df.columns:
    df[resource_column_name] = df[resource_column_name].apply(lambda x: str(x).split("/")[-1])

# Step 5: Drop unwanted columns
df.drop(columns=[col for col in columns_to_remove if col in df.columns], inplace=True)

# Step 6: Add required empty columns if missing
for col in columns_to_add:
    if col not in df.columns:
        df[col] = ""

# Step 7: Subscription mapping
df_sub = pd.read_excel(subscription_file)
df_sub.columns = df_sub.columns.str.strip()
validate_required_columns(df_sub, ["Subscription ID", "Environment", primary_contact_column], "Subscription File")
df_sub["Subscription ID"] = df_sub["Subscription ID"].astype(str).str.strip()
df["Subscription ID"] = df["Subscription ID"].astype(str).str.strip()
df.drop(columns=[col for col in df_sub.columns if col in df.columns and col not in ["Subscription ID", "Subscription Name"]], inplace=True, errors="ignore")

unmatched_subs = sorted(set(df["Subscription ID"]) - set(df_sub["Subscription ID"]))
if unmatched_subs:
    with open("unmatched_subscription_id.txt", "w") as f:
        f.writelines(f"{id}\n" for id in unmatched_subs)
    print(f"❌ {len(unmatched_subs)} unmatched Subscription ID entries logged.")
else:
    print("✅ All Subscription IDs matched.")

df = df.merge(df_sub[["Subscription ID", "Environment", primary_contact_column]], on="Subscription ID", how="left")
df["Environment"] = df["Environment"].fillna("Environment not available")
df[primary_contact_column] = df[primary_contact_column].fillna("Primary contact not available")

# Step 8: Remediation mapping
df_remed = pd.read_excel(remediation_file)
df_remed.columns = df_remed.columns.str.strip()
validate_required_columns(df_remed, ["Policy ID", "Policy Statement", "Policy Remediation"], "Remediation File")
df_remed["Policy ID"] = df_remed["Policy ID"].astype(str).str.strip()
df["Policy ID"] = df["Policy ID"].astype(str).str.strip()
df.drop(columns=["Policy Statement", "Policy Remediation"], inplace=True, errors="ignore")

unmatched_policies = sorted(set(df["Policy ID"]) - set(df_remed["Policy ID"]))
if unmatched_policies:
    with open("unmatched_policy_id.txt", "w") as f:
        f.writelines(f"{id}\n" for id in unmatched_policies)
    print(f"❌ {len(unmatched_policies)} unmatched Policy ID entries logged.")
else:
    print("✅ All Policy IDs matched.")

df = df.merge(df_remed[["Policy ID", "Policy Statement", "Policy Remediation"]], on="Policy ID", how="left")
df["Description"] = df["Policy Statement"].fillna("Policy details not available")
df["Remediation Steps"] = df["Policy Remediation"].fillna("Remediation steps not available")

# Step 9: Contact Hierarchy mapping
df_contact = pd.read_excel(ownership_file)
df_contact.columns = df_contact.columns.str.strip()
validate_required_columns(df_contact, [primary_contact_column] + manager_columns, "Ownership File")
df.drop(columns=[col for col in df_contact.columns if col in df.columns and col != primary_contact_column], inplace=True, errors="ignore")
df = df.merge(df_contact[[primary_contact_column] + manager_columns], on=primary_contact_column, how="left")
print("✅ Mapped Manager Hierarchy and BU.")

# Step 10: Reorder columns
ordered = [col for col in final_columns if col in df.columns]
df = df[ordered + [col for col in df.columns if col not in ordered]]

# Step 11: Write to Excel with formatting
print("\n💾 Saving Excel file with formatting...")
df.to_excel(output_excel, index=False)

wb = load_workbook(output_excel)
ws = wb.active

align = Alignment(horizontal="left", vertical="top", wrap_text=True)
header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
header_font = Font(bold=True)

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = align

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

ws.freeze_panes = "A2"
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

wb.save(output_excel)
print(f"✅ Final file saved as: {output_excel}")

# Final execution time
total_time = time.time() - start_time
print(f"⏱️ Total time: {format_duration(total_time)}")

# Final validation of required columns
missing_final = [col for col in final_columns if col not in df.columns]
if missing_final:
    print(f"❌ Final output is missing columns: {missing_final}")
else:
    print("✅ Final output contains all required columns.")
