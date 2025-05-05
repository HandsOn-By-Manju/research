import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time

# === File paths ===
input_csv = "input_file.csv"
output_excel = "output_file.xlsx"
anex_file = "Anex.xlsx"

# === Columns to remove or add ===
columns_to_remove = [
    "DummyColumn1", "DummyColumn2", "DummyColumn3",
    "DummyColumn4", "DummyColumn5", "DummyColumn6",
    "DummyColumn7", "DummyColumn8", "DummyColumn9"
]
columns_to_add = ["Col1", "Col2", "Col3", "Col4", "Col5", "Col6", "Col7", "Col8"]

start_time = time.time()
print("\n🚀 Starting Excel processing...")

# === Step 1: Load input CSV ===
df = pd.read_csv(input_csv)
print("✅ Loaded input CSV.")

# === Step 2: Extract 'Subscription ID' and 'Subscription Name' from 'Account' column ===
if 'Account' in df.columns:
    df['Subscription ID'] = df['Account'].str.extract(r'^(\S+)\s*\(')[0].str.replace(r'\s+', '', regex=True)
    df['Subscription Name'] = df['Account'].str.extract(r'\((.*?)\)')[0].str.replace(r'\s+', '', regex=True)

# === Step 3: Clean 'Resource ID' to just filename ===
if 'Resource ID' in df.columns:
    df['Resource ID'] = df['Resource ID'].apply(lambda x: str(x).split('/')[-1])

# === Step 4: Drop unwanted columns ===
df.drop(columns=[col for col in columns_to_remove if col in df.columns], inplace=True)

# === Step 5: Add new blank columns ===
for col in columns_to_add:
    df[col] = ''

# === Step 6: Merge Policy info from Anex1 ===
anex1 = pd.read_excel(anex_file, sheet_name="Anex1")
df = df.merge(anex1, on="Policy ID", how="left", suffixes=('', '_anex1'))
df['Col1'] = df['Policy Statement_anex1'].fillna("Policy details doesn't exist")
df['Col2'] = df['Policy Remediation_anex1'].fillna("Policy details doesn't exist")

# Log unmatched policy IDs
unmatched_policy_ids = df[df['Policy Statement_anex1'].isna() & df['Policy Remediation_anex1'].isna()]['Policy ID'].dropna().unique()
if len(unmatched_policy_ids) > 0:
    with open("unmatched_policy_ids.txt", "w") as f:
        for pid in unmatched_policy_ids:
            f.write(f"{pid}\n")
df.drop(columns=['Policy Statement_anex1', 'Policy Remediation_anex1'], inplace=True)

# === Step 7: Merge Subscription info from Anex2 ===
anex2 = pd.read_excel(anex_file, sheet_name="Anex2")
df = df.merge(anex2, on="Subscription ID", how="left", suffixes=('', '_anex2'))
df['Col3'] = df['Environment_anex2'].fillna("Environment/contact info not found")
df['Col4'] = df['Primary Contact_anex2'].fillna("Environment/contact info not found")

# Log unmatched subscription IDs
unmatched_subs = df[df['Environment_anex2'].isna() & df['Primary Contact_anex2'].isna()]['Subscription ID'].dropna().unique()
if len(unmatched_subs) > 0:
    with open("unmatched_subscription_ids.txt", "w") as f:
        for sid in unmatched_subs:
            f.write(f"{sid}\n")
df.drop(columns=['Environment_anex2', 'Primary Contact_anex2'], inplace=True)

# === Step 8: Merge Contact info from Anex3 ===
anex3 = pd.read_excel(anex_file, sheet_name="Anex3")
df = df.merge(anex3, on="Contact", how="left", suffixes=('', '_anex3'))
df['Col5'] = df['M1'].fillna("Data not available")
df['Col6'] = df['M2'].fillna("Data not available")
df['Col7'] = df['M3'].fillna("Data not available")
df['Col8'] = df['M4'].fillna("Data not available")

# Log unmatched contacts
unmatched_contacts = df[df[['M1', 'M2', 'M3', 'M4']].isna().all(axis=1)]['Contact'].dropna().unique()
if len(unmatched_contacts) > 0:
    with open("unmatched_anex3_contacts.txt", "w") as f:
        for contact in unmatched_contacts:
            f.write(f"{contact}\n")
df.drop(columns=['M1', 'M2', 'M3', 'M4'], inplace=True)

# === Step 9: Save to Excel ===
df.to_excel(output_excel, index=False)
print(f"💾 Excel saved to: {output_excel}")

# === Step 10: Format Excel file ===
wb = load_workbook(output_excel)
ws = wb.active
ws.title = "Issues"
alignment = Alignment(horizontal='left', vertical='top')
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = alignment
wb.save(output_excel)

# === Done ===
print(f"✅ Script completed in {time.time() - start_time:.2f} seconds.")
