import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Step 1: Read CSV file
csv_file_path = 'input_file.csv'  # Replace with your actual file path
df = pd.read_csv(csv_file_path)

# Step 2: Show original columns
print("Original Columns:")
print(df.columns.tolist())

# Step 3: Extract 'ID' and 'Name' from 'Details'
if 'Details' in df.columns:
    df['ID'] = df['Details'].str.extract(r'^\s*(\S+)\s*\(')[0].str.replace(r'\s+', '', regex=True)
    df['Name'] = df['Details'].str.extract(r'\((.*?)\)')[0].str.replace(r'\s+', '', regex=True)

# Step 4: Format 'Res_ID' to keep only filename from forward-slash path
if 'Res_ID' in df.columns:
    df['Res_ID'] = df['Res_ID'].apply(lambda x: str(x).split('/')[-1])

# Step 5: Define columns to remove (edit this list)
columns_to_remove = [
    'DummyColumn1',
    'DummyColumn2',
    'DummyColumn3',
    'DummyColumn4',
    'DummyColumn5',
    'DummyColumn6',
    'DummyColumn7',
    'DummyColumn8',
    'DummyColumn9'
]

# Drop only existing columns from the list
existing_cols_to_remove = [col for col in columns_to_remove if col in df.columns]
df.drop(columns=existing_cols_to_remove, inplace=True)

# Step 6: Add new columns 'Description' and 'Remediation'
df['Description'] = ''
df['Remediation'] = ''

# Step 6.1: Add 8 empty columns named Col1 to Col8
new_columns = ['Col1', 'Col2', 'Col3', 'Col4', 'Col5', 'Col6', 'Col7', 'Col8']
for col in new_columns:
    df[col] = ''

# Step 7: Show updated columns
print("\nUpdated Columns:")
print(df.columns.tolist())

# Step 8: Save to Excel
excel_file_path = 'output_file.xlsx'
df.to_excel(excel_file_path, index=False)

# Step 9: Apply top-left alignment and rename sheet
wb = load_workbook(excel_file_path)
ws = wb.active
ws.title = "Issues"  # Rename the sheet to 'Issues'

alignment = Alignment(horizontal='left', vertical='top')
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = alignment

wb.save(excel_file_path)
print(f"\n✅ Excel file saved successfully with formatting and sheet renamed: {excel_file_path}")
