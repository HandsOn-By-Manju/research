import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Configurable input Excel file name
input_file = 'your_input_file.xlsx'  # <-- Replace with your actual file name
output_file = 'Subscription_Details_Master_Data.xlsx'

try:
    # Check if file exists
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Input file '{input_file}' not found.")

    # Read the Excel file
    df = pd.read_excel(input_file)

    # List all columns
    print("Columns in the Excel file:")
    print(df.columns.tolist())

    # Ensure required columns exist
    required_cols = ['id', 'name']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise KeyError(f"Missing required column(s): {', '.join(missing_cols)}")

    # Remove '/subscription/' prefix from 'id'
    df['id'] = df['id'].astype(str).str.replace('/subscription/', '', regex=False)

    # Rename columns
    df.rename(columns={
        'id': 'Subscription ID',
        'name': 'Subscription Name',
        'PrimaryContact': 'Primary Contact'
    }, inplace=True)

    # Add missing columns
    if 'Environment' not in df.columns:
        df['Environment'] = ''
    if 'BU' not in df.columns:
        df['BU'] = ''

    # Assign BU values
    df['BU'] = df['Subscription Name'].astype(str).apply(
        lambda x: 'Commerce' if 'commerce' in x.lower() else '')

    if 'Primary Contact' in df.columns:
        df.loc[df['Primary Contact'].astype(str).str.lower() == 'test@test.com', 'BU'] = 'Alpha'

    df['BU'] = df['BU'].replace('', 'Beta')

    # Assign Environment values
    if 'ManagementGroup' in df.columns:
        df['Environment'] = df['ManagementGroup'].astype(str).apply(
            lambda x: 'Production' if x.startswith('Prod-') else
                      'Non-Production' if x.startswith('Non-Prod-') else
                      'Development' if x.startswith('Dev-') else 'Production'
        )
    else:
        df['Environment'] = 'Production'

    # Save to Excel
    df.to_excel(output_file, index=False)

    # Apply top and left alignment using openpyxl
    wb = load_workbook(output_file)
    ws = wb.active
    align_style = Alignment(vertical='top', horizontal='left')

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align_style

    wb.save(output_file)
    print(f"\n✅ Processed file saved with alignment as: {output_file}")

except FileNotFoundError as e:
    print(f"❌ Error: {e}")
except KeyError as e:
    print(f"❌ Error: {e}")
except Exception as e:
    print(f"❌ An unexpected error occurred: {e}")
