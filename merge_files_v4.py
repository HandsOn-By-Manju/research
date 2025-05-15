import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# === START TIMER ===
start_time = time.time()

# === CONFIGURATION ===
folder_path = os.getcwd()
output_file = "merged_output.xlsx"
log_file = os.path.join(folder_path, "merge_log.txt")

print(f"üîç Current folder being scanned: {folder_path}")

# === SAFETY CHECK ===
if not os.path.isdir(folder_path):
    print(f"‚ùå The folder does not exist: {folder_path}")
    exit()

# === INITIATE LOG ===
log_lines = []

# === STEP 1: Get All Excel Files ===
print("üì• Searching for Excel files...")
excel_files = [f for f in os.listdir(folder_path) if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
log_lines.append(f"Number of Excel files found: {len(excel_files)}")
print(f"‚úÖ Found {len(excel_files)} Excel files.")

if len(excel_files) == 0:
    log_lines.append("No Excel files found.")
    with open(log_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(log_lines))
    print("‚ùå No Excel files found in current folder.")
    exit()

# === STEP 2: Validate Column Names ===
print("üîé Checking column structure of files...")
first_file = os.path.join(folder_path, excel_files[0])
first_df = pd.read_excel(first_file)
expected_columns = list(first_df.columns)
log_lines.append(f"Expected columns: {expected_columns}")
print(f"üìå Expected columns: {expected_columns}")

all_dataframes = []
row_counts = {}

for file in excel_files:
    print(f"üîç Reading file: {file}")
    path = os.path.join(folder_path, file)
    df = pd.read_excel(path)

    if list(df.columns) != expected_columns:
        log_lines.append(f"‚ùå Column mismatch in file: {file}")
        log_lines.append(f"Found columns: {list(df.columns)}")
        with open(log_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(log_lines))
        print(f"‚ùå Column mismatch found in '{file}'. Check merge_log.txt for details.")
        exit()

    row_counts[file] = len(df)
    all_dataframes.append(df)
    print(f"‚úÖ File '{file}' loaded with {len(df)} rows.")

# === STEP 3: Merge Files ===
print("üîß Merging all data...")
merged_df = pd.concat(all_dataframes, ignore_index=True)
output_path = os.path.join(folder_path, output_file)
merged_df.to_excel(output_path, index=False)
print(f"‚úÖ Merged data written to '{output_file}'")

# === STEP 4: Excel Formatting ===
print("üé® Applying Excel formatting...")
wb = load_workbook(output_path)
ws = wb.active

header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill

for col in ws.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_length + 2

ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = "A2"
wb.save(output_path)
print("‚úÖ Excel formatting complete.")

# === STEP 5: Write Log File ===
log_lines.append("\nRow counts per file:")
for file, count in row_counts.items():
    log_lines.append(f"{file}: {count} rows")

log_lines.append(f"\nTotal rows in merged file: {len(merged_df)}")

# === STEP 6: Show Time Taken in Smart Format ===
end_time = time.time()
elapsed = int(end_time - start_time)

def format_time(seconds):
    if seconds < 60:
        return f"{seconds} seconds"
    elif seconds < 3600:
        minutes = seconds // 60
        sec = seconds % 60
        return f"{minutes} minutes {sec} seconds"
    else:
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        sec = seconds % 60
        return f"{hours} hours {minutes} minutes {sec} seconds"

formatted_time = format_time(elapsed)
log_lines.append(f"\n‚è±Ô∏è Time taken: {formatted_time}")

with open(log_file, "w", encoding="utf-8") as f:
    f.write("\n".join(log_lines))

# === FINAL OUTPUT ===
print("üìÑ Log written to merge_log.txt")
print("\nüìä Summary:")
for file, count in row_counts.items():
    print(f"   - {file}: {count} rows")
print(f"\nüìà Total rows merged: {len(merged_df)}")
print(f"‚è±Ô∏è Time taken to complete: {formatted_time}")
print(f"\n‚úÖ Merged Excel file: {output_path}")
print(f"üìù Log file created at: {log_file}")
