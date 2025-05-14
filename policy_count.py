import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# 1. Define Excel file paths for each Business Unit (BU)
files = {
    'BU1': 'BU1.xlsx',
    'BU2': 'BU2.xlsx',
    'BU3': 'BU3.xlsx',
    'BU4': 'BU4.xlsx',
}

# Store grouped data for each BU
bu_counts = {}
# Store summary (BU name + unique Policy ID count)
summary_data = []

# 2. Process each BU file
for bu, file in files.items():
    df = pd.read_excel(file)

    # Clean: trim spaces in important columns
    df['Policy ID'] = df['Policy ID'].astype(str).str.strip()
    df['Policy Name'] = df['Policy Name'].astype(str).str.strip()
    df['Severity'] = df['Severity'].astype(str).str.strip()
    df['Service'] = df['Service'].astype(str).str.strip()

    # Drop rows where Policy ID is missing
    df.dropna(subset=['Policy ID'], inplace=True)

    # Add to summary: count unique Policy IDs in this BU
    summary_data.append({'BU': bu, 'Unique Policy Count': df['Policy ID'].nunique()})

    # Group by 4 key fields and count rows
    counts = df.groupby(['Policy ID', 'Policy Name', 'Severity', 'Service']).size().reset_index(name=bu)
    bu_counts[bu] = counts

# 3. Merge data from all BUs
merged_df = bu_counts['BU1']
for bu in ['BU2', 'BU3', 'BU4']:
    merged_df = pd.merge(merged_df, bu_counts[bu], on=['Policy ID', 'Policy Name', 'Severity', 'Service'], how='outer')

# 4. Handle missing BU columns
for bu in ['BU1', 'BU2', 'BU3', 'BU4']:
    if bu not in merged_df:
        merged_df[bu] = 0
merged_df.fillna(0, inplace=True)

# 5. Convert BU count columns to integers
for bu in ['BU1', 'BU2', 'BU3', 'BU4']:
    merged_df[bu] = merged_df[bu].astype(int)

# 6. Add a Total Count column (sum of all BU counts)
merged_df['Total Count'] = merged_df[['BU1', 'BU2', 'BU3', 'BU4']].sum(axis=1)

# 7. Save merged data and summary to Excel (2 sheets)
output_file = "Combined_Policy_Counts_With_Summary.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    merged_df.to_excel(writer, index=False, sheet_name='Policy Data')
    pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name='Summary')

# 8. Excel formatting using openpyxl
wb = load_workbook(output_file)

# === Format Policy Data Sheet ===
ws = wb['Policy Data']
ws.freeze_panes = 'A2'  # Freeze header row
header_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')  # Sky blue fill

for col_num, cell in enumerate(ws[1], 1):
    cell.fill = header_fill
    # Auto-fit column width
    max_length = max(
        len(str(cell.value)) if cell.value else 0,
        max((len(str(ws.cell(row=i, column=col_num).value)) for i in range(2, ws.max_row + 1)), default=0)
    )
    ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

ws.auto_filter.ref = ws.dimensions  # Enable filter on all columns

# === Format Summary Sheet ===
summary_ws = wb['Summary']
for col_num, cell in enumerate(summary_ws[1], 1):
    cell.fill = header_fill
    max_length = max(
        len(str(cell.value)) if cell.value else 0,
        max((len(str(summary_ws.cell(row=i, column=col_num).value)) for i in range(2, summary_ws.max_row + 1)), default=0)
    )
    summary_ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

# 9. Save final Excel file
wb.save(output_file)

print(f"✅ Final Excel file with Policy Data and Summary saved as '{output_file}'")
