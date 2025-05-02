import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Step 1: Read CSV file
csv_file_path = 'input_file.csv'  # Replace with your actual file path
df = pd.read_csv(csv_file_path)

# Step 2: Show original columns
print("Original Columns:")
print(df.columns.tolist())

# Step 3: Extract 'ID' and 'Name' from 'Details'
if 'Details' in df.columns:
    df['ID'] = df['Details'].str.extract(r'^\s*(\S+)\s*\(')[0].str.replace(r'\s+', '', regex=True)
    df['Name'] = df['Details'].str.extract(r'\((.*?)\)')[0].str.replace(r'\s+', '', regex=True)

# Step 4: Define columns to remove
columns_to_remove = [
    'DummyColumn1',
    'DummyColumn2',
    'DummyColumn3',
    'DummyColumn4',
    'DummyColumn5',
    'DummyColumn6',
    'DummyColumn7',
    'DummyColumn8',
    'DummyColumn9'
]

# Drop only columns that exist
existing_cols_to_remove = [col for col in columns_to_remove if col in df.columns]
df.drop(columns=existing_cols_to_remove, inplace=True)

# Step 5: Show updated columns
print("\nUpdated Columns:")
print(df.columns.tolist())

# Step 6: Save to Excel
excel_file_path = 'output_file.xlsx'
df.to_excel(excel_file_path, index=False)

# Step 7: Apply top-left alignment only
wb = load_workbook(excel_file_path)
ws = wb.active
alignment = Alignment(horizontal='left', vertical='top')

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = alignment

wb.save(excel_file_path)
print(f"\nExcel file saved with top-left alignment and selected columns removed: {excel_file_path}")
