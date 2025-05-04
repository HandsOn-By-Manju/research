import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import json
import time

# Load configuration
with open('config.json') as f:
    config = json.load(f)

input_csv = config["input_csv"]
output_excel = config["output_excel"]
anex_file = config["anex_file"]
columns_to_remove = config["columns_to_remove"]
columns_to_add = config["columns_to_add"]
mappings = config["mappings"]

start_time = time.time()
print("\n🚀 Starting the Excel processing script...")

# Step 1: Read CSV
step_start = time.time()
print(f"📄 Reading input CSV: {input_csv}")
df = pd.read_csv(input_csv)
print(f"✅ Loaded CSV in {time.time() - step_start:.2f} seconds. Columns: {df.columns.tolist()}")

# Step 2: Extract ID and Name
step_start = time.time()
if 'Details' in df.columns:
    print("🔧 Extracting ID and Name from 'Details'")
    df['ID'] = df['Details'].str.extract(r'^\s*(\S+)\s*\(')[0].str.replace(r'\s+', '', regex=True)
    df['Name'] = df['Details'].str.extract(r'\((.*?)\)')[0].str.replace(r'\s+', '', regex=True)
    print(f"✅ ID and Name extracted in {time.time() - step_start:.2f} seconds.")

# Step 3: Clean Res_ID
step_start = time.time()
if 'Res_ID' in df.columns:
    print("🔧 Cleaning 'Res_ID'...")
    df['Res_ID'] = df['Res_ID'].apply(lambda x: str(x).split('/')[-1])
    print(f"✅ Res_ID cleaned in {time.time() - step_start:.2f} seconds.")

# Step 4: Drop dummy columns
step_start = time.time()
print("🧹 Dropping dummy columns if any...")
existing_cols = [col for col in columns_to_remove if col in df.columns]
df.drop(columns=existing_cols, inplace=True)
print(f"✅ Dropped: {existing_cols if existing_cols else 'None'} in {time.time() - step_start:.2f} seconds.")

# Step 5: Add blank columns
step_start = time.time()
print(f"➕ Adding columns: {columns_to_add}")
for col in columns_to_add:
    df[col] = ''
print(f"✅ Columns added in {time.time() - step_start:.2f} seconds.")

# Step 6: Process each mapping
for mapping in mappings:
    step_start = time.time()
    sheet_name = mapping['sheet']
    key_column = mapping['key']
    source_to_target = mapping['columns']
    not_found_message = mapping['not_found_message']
    unmatched_log_file = mapping['unmatched_log']

    print(f"\n📄 Reading sheet '{sheet_name}' from {anex_file} for mapping on '{key_column}'...")
    map_df = pd.read_excel(anex_file, sheet_name=sheet_name)
    relevant_columns = [key_column] + list(source_to_target.keys())
    map_df = map_df[relevant_columns]

    print(f"🔗 Merging on '{key_column}'...")
    df = df.merge(map_df, on=key_column, how='left')

    # Fill values and check unmatched
    for src_col, target_col in source_to_target.items():
        df[target_col] = df[src_col].fillna(not_found_message)

    unmatched = df[list(source_to_target.keys())].isna().all(axis=1)
    unmatched_values = df.loc[unmatched, key_column].dropna().unique()

    if len(unmatched_values) > 0:
        print(f"⚠️ Unmatched values for '{key_column}' in sheet '{sheet_name}':")
        for val in unmatched_values:
            print(f"- {val}")
        with open(unmatched_log_file, "w") as f:
            for val in unmatched_values:
                f.write(f"{val}\n")
        print(f"📁 Unmatched values saved to '{unmatched_log_file}'")
    else:
        print(f"✅ All values matched for '{key_column}'")

    df.drop(columns=list(source_to_target.keys()), inplace=True, errors='ignore')
    print(f"✅ Mapping from '{sheet_name}' complete in {time.time() - step_start:.2f} seconds.")

# Step 7: Save to Excel
step_start = time.time()
print(f"\n💾 Writing output Excel to: {output_excel}")
df.to_excel(output_excel, index=False)
print(f"✅ Excel file written in {time.time() - step_start:.2f} seconds.")

# Step 8: Format Excel
step_start = time.time()
print("🎨 Formatting Excel file...")
wb = load_workbook(output_excel)
ws = wb.active
ws.title = "Issues"
alignment = Alignment(horizontal='left', vertical='top')
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = alignment
wb.save(output_excel)
print(f"✅ Formatting complete in {time.time() - step_start:.2f} seconds.")

# Done
total_time = time.time() - start_time
print(f"\n✅ Script completed in {total_time:.2f} seconds. Output: {output_excel}")
