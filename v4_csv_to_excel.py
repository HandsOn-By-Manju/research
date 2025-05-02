import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Step 1: Read the CSV file
csv_file_path = 'input_file.csv'  # Replace with your file name
df = pd.read_csv(csv_file_path)

# Step 2: List original columns
print("Original Columns:")
print(df.columns.tolist())

# Step 3: Process 'Details' column
if 'Details' in df.columns:
    df['ID'] = df['Details'].str.extract(r'^\s*(\S+)\s*\(')[0].str.replace(r'\s+', '', regex=True)
    df['Name'] = df['Details'].str.extract(r'\((.*?)\)')[0].str.replace(r'\s+', '', regex=True)

# Step 4: Print updated column list
print("\nUpdated Columns:")
print(df.columns.tolist())

# Step 5: Save to Excel
excel_file_path = 'output_file.xlsx'
df.to_excel(excel_file_path, index=False)

# Step 6: Apply top-left alignment using openpyxl
wb = load_workbook(excel_file_path)
ws = wb.active

alignment = Alignment(horizontal='left', vertical='top')

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = alignment

wb.save(excel_file_path)
print(f"\nExcel file saved with top-left alignment: {excel_file_path}")
