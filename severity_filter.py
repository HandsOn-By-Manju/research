import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- Config ---
input_file = "input.xlsx"  # Change to your actual file path
output_file = "filtered_critical.xlsx"
sheet_name = "Sheet1"
column_name = "Severity"
filter_value = "Critical"

# --- Step 1: Read Excel and Filter Data ---
df = pd.read_excel(input_file, sheet_name=sheet_name)
df_filtered = df[df[column_name].str.strip().str.lower() == filter_value.lower()]

# --- Step 2: Write Filtered Data to New Excel File ---
df_filtered.to_excel(output_file, index=False)

# --- Step 3: Apply Formatting using openpyxl ---
wb = load_workbook(output_file)
ws = wb.active

# Freeze top row
ws.freeze_panes = "A2"

# Apply formatting to header
header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

# Format header
for cell in ws[1]:
    cell.fill = header_fill
    cell.alignment = alignment

# Format all data cells
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = alignment

# Auto-fit column width
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width

# Save the final formatted file
wb.save(output_file)

print(f"Filtered rows with '{filter_value}' severity saved and formatted in '{output_file}'")
