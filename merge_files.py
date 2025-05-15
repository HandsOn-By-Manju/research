import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# === CONFIGURATION ===
folder_path = 'your_folder_path_here'  # 🔁 Change this to your folder path
output_file = 'merged_output.xlsx'
log_file = os.path.join(folder_path, 'merge_log.txt')

# === INITIATE LOG ===
log_lines = []

# === STEP 1: Get All Excel Files ===
excel_files = [f for f in os.listdir(folder_path) if f.endswith(('.xlsx', '.xls'))]
log_lines.append(f"Number of Excel files found: {len(excel_files)}")

if len(excel_files) == 0:
    log_lines.append("No Excel files found.")
    with open(log_file, 'w') as f:
        f.write('\n'.join(log_lines))
    exit()

# === STEP 2: Check If All Files Have Same Columns ===
first_df = pd.read_excel(os.path.join(folder_path, excel_files[0]))
expected_columns = list(first_df.columns)
log_lines.append(f"Expected columns: {expected_columns}")

all_dataframes = []
row_counts = {}

for file in excel_files:
    path = os.path.join(folder_path, file)
    df = pd.read_excel(path)

    if list(df.columns) != expected_columns:
        log_lines.append(f"❌ Column mismatch in file: {file}")
        log_lines.append(f"Found columns: {list(df.columns)}")
        with open(log_file, 'w') as f:
            f.write('\n'.join(log_lines))
        print(f"Column mismatch found. Check {log_file} for details.")
        exit()

    row_counts[file] = len(df)
    all_dataframes.append(df)

# === STEP 3: Merge All DataFrames ===
merged_df = pd.concat(all_dataframes, ignore_index=True)
output_path = os.path.join(folder_path, output_file)
merged_df.to_excel(output_path, index=False)

# === STEP 4: Apply Excel Formatting ===
wb = load_workbook(output_path)
ws = wb.active

# Make header bold and blue
header_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill

# Auto-fit column width
for col in ws.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_length + 2

# Add autofilter and freeze header row
ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = 'A2'
wb.save(output_path)

# === STEP 5: Write Log File ===
log_lines.append("\nRow counts per file:")
for file, count in row_counts.items():
    log_lines.append(f"{file}: {count} rows")

log_lines.append(f"\nTotal rows in merged file: {len(merged_df)}")

with open(log_file, 'w') as f:
    f.write('\n'.join(log_lines))

# === STEP 6: Final Message ===
print("\n".join(log_lines))
print(f"\n✅ Merged Excel file saved at: {output_path}")
print(f"📄 Log file created at: {log_file}")
